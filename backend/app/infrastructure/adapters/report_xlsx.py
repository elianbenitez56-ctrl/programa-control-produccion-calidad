"""Adaptador XLSX del Módulo Reportes (openpyxl).

Implementa el puerto `ReportBuilder`. Es infraestructura pura: no contiene
lógica de negocio ni datos; los casos de uso del Módulo 12 la alimentarán.
"""
from collections.abc import Iterable
from io import BytesIO
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from app.application.ports.report_ports import ReportBuilder

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SUBTITLE_FONT = Font(italic=True, color="64748B")


class XlsxReportBuilder(ReportBuilder):
    """Construye un libro Excel con una hoja de datos estructurados."""

    def __init__(self, sheet_name: str = "Reporte") -> None:
        self._workbook = Workbook()
        self._sheet = self._workbook.active
        self._sheet.title = sheet_name
        self._row = 1

    def add_title(self, title: str, subtitle: str | None = None) -> None:
        self._sheet.cell(self._row, 1, title).font = Font(bold=True, size=14)
        self._row += 1
        if subtitle:
            self._sheet.cell(self._row, 1, subtitle).font = _SUBTITLE_FONT
            self._row += 1
        self._row += 1

    def add_section(self, titulo: str, campos: dict[str, Any]) -> None:
        self._sheet.cell(self._row, 1, titulo).font = Font(bold=True, size=11)
        self._row += 1
        for clave, valor in campos.items():
            celda = self._sheet.cell(self._row, 1, str(clave))
            celda.font = Font(bold=True)
            self._sheet.cell(self._row, 2, str(valor))
            self._row += 1
        self._row += 1

    def add_table(self, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
        rows_list = [list(row) for row in rows]
        for col, header in enumerate(headers, start=1):
            celda = self._sheet.cell(self._row, col, header)
            celda.fill = _HEADER_FILL
            celda.font = _HEADER_FONT
        self._row += 1
        for row in rows_list:
            for col, value in enumerate(row, start=1):
                self._sheet.cell(self._row, col, value)
            self._row += 1
        self._row += 1

        for col in range(1, len(headers) + 1):
            max_len = len(str(headers[col - 1]))
            for row in rows_list:
                if len(row) >= col:
                    max_len = max(max_len, len(str(row[col - 1])))
            self._sheet.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    def to_bytes(self) -> bytes:
        buffer = BytesIO()
        self._workbook.save(buffer)
        return buffer.getvalue()
